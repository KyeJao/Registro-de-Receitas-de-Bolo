from flask import Flask, render_template, request, redirect
from openpyxl import Workbook, load_workbook
import os

bolo = Flask(__name__)

ARQUIVO = 'receitas.xlsx'

if not os.path.exists(ARQUIVO):
    wb = Workbook()
    ws = wb.active

    ws.append(["Nome do Bolo", "URL", "1° ingrediente","2° ingrediente","3° ingrediente","4° ingrediente","5° ingrediente"])
    wb.save(ARQUIVO)

@bolo.route('/')
def index():
    return render_template('index.html')

@bolo.route('/salvar', methods=["POST"])
def salvar():
    nome = request.form['nomeBolo']
    imagemBol = request.form['imgBolo']
    ing_um = request.form['ingredientes']
    ing_dois = request.form['ingredientes2']
    ing_tres = request.form['ingredientes3']
    ing_quack = request.form['ingredientes4']
    ing_quin = request.form['ingredientes5']
    wb =  load_workbook(ARQUIVO)
    ws = wb.active
    ws.append([nome,imagemBol,ing_um,ing_dois,ing_tres,ing_quack,ing_quin])
    wb.save(ARQUIVO)
    return redirect('/tabela')
    
@bolo.route('/tabela')
def tabela():
    wb = load_workbook(ARQUIVO)
    ws = wb.active
    dados = list(ws.iter_rows(min_row=2,values_only=True))
    return render_template('tabela.html',dados = dados)

if __name__ == '__main__':
    bolo.run(debug=True)